import csv
import io
from typing import Annotated

import httpx
from fastapi import APIRouter, Body, Depends, File, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.v1.dependencies import get_catalog_controller
from app.api.v1.route_factory import register_catalog_routes
from app.controllers.catalog_controller import CatalogController
from app.core.database import get_db
from app.core.security import CurrentUser, get_current_user
from app.helpers.response import accepted
from app.services.org_setup_service import OrgSetupService

router = APIRouter()

# Bulk-upload routes are handled below as direct file-upload endpoints.
ENDPOINTS = [
    ("GET",  "/progress",          "org_setup_progress_get",          "Get setup progress"),
    ("POST", "/step1",             "org_setup_step1_save",            "Save step 1 – org details"),
    ("GET",  "/step1",             "org_setup_step1_get",             "Get step 1 – org details"),
    ("POST", "/step2",             "org_setup_step2_save",            "Save step 2 – compliance"),
    ("GET",  "/step2",             "org_setup_step2_get",             "Get step 2 – compliance"),
    ("POST", "/step3/site",        "org_setup_step3_site_create",     "Create a site"),
    ("GET",  "/step3/sites",       "org_setup_step3_sites_list",      "List sites"),
    ("POST", "/step4/user",        "org_setup_step4_user_create",     "Create a user"),
    ("GET",  "/step4/users",       "org_setup_step4_users_list",      "List users"),
    ("POST", "/step5",             "org_setup_step5_save",            "Save step 5 – workflow config"),
    ("GET",  "/step5",             "org_setup_step5_get",             "Get step 5 – workflow config"),
    ("POST", "/step6/upload",      "org_setup_step6_upload",          "Upload knowledge document"),
    ("GET",  "/step6/documents",   "org_setup_step6_documents_list",  "List knowledge documents"),
    ("POST", "/step6a/import",     "org_setup_step6a_import",         "Import historical data"),
    ("GET",  "/step6a/imports",    "org_setup_step6a_imports_list",   "List data imports"),
    ("POST", "/step7",             "org_setup_step7_save",            "Save step 7 – AI config"),
    ("GET",  "/step7",             "org_setup_step7_get",             "Get step 7 – AI config"),
    ("POST", "/activate",          "org_setup_activate",              "Activate organisation (step 8)"),
]

register_catalog_routes(router, "org_setup", ENDPOINTS)

# ── File-parsing helper ────────────────────────────────────────────────────────

_MODULE_PREFERRED_SHEETS: dict[str, list[str]] = {
    "training":         ["Training Module Library", "Employee Training Records", "Assessment Records"],
    "training_records": ["Training Module Library", "Employee Training Records", "Assessment Records"],
    "incidents":        ["Incident Records", "Incidents", "Incident Log"],
    "audits":           ["Audit Reports", "Audits", "Audit Log"],
    "risk":             ["Risk Register", "Risk Assessments", "Critical Risks", "High Risks", "Medium Risks", "Low Risks"],
    "risk_assessments": ["Risk Register", "Risk Assessments", "Critical Risks", "High Risks", "Medium Risks", "Low Risks"],
    "vendors":          ["Contractor Records", "Vendors", "Contractors"],
    "assets":           ["Asset Records", "Assets", "Equipment"],
}

# For these modules, combine all matching sheets into one list of rows
_MODULE_COMBINE_SHEETS: dict[str, list[str]] = {
    "risk":             ["Critical Risks", "High Risks", "Medium Risks", "Low Risks"],
    "risk_assessments": ["Critical Risks", "High Risks", "Medium Risks", "Low Risks"],
}


def _parse_excel_sheet(ws) -> list[dict]:
    """Parse one openpyxl worksheet into row dicts.

    Scans the first 5 rows to find the real header row — skips leading title/
    empty rows that are common in client-supplied Excel files.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Find the first row with 2+ non-empty cells — that's the header row
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        filled = sum(1 for v in row if v is not None and str(v).strip())
        if filled >= 2:
            header_idx = i
            break
    headers = [str(h or "").strip() for h in rows[header_idx]]
    result = []
    for row in rows[header_idx + 1:]:
        if any(v is not None and str(v).strip() for v in row):
            result.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
    return result


def _parse_file(content: bytes, filename: str, module: str = "") -> list[dict]:
    """Parse CSV or Excel bytes into a list of row dicts."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        text = content.decode("utf-8-sig")
        # Auto-detect delimiter: tab if more tabs than commas on first line, else comma
        first_line = text.split("\n")[0] if text else ""
        delimiter = "\t" if first_line.count("\t") > first_line.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        return [dict(row) for row in reader]
    # Excel (.xlsx / .xls)
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl not installed – please upload a CSV file instead")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open Excel file: {exc}") from exc

    try:
        # Modules that split data across multiple sheets — combine them all
        combine_names = _MODULE_COMBINE_SHEETS.get(module, [])
        if combine_names:
            combined: list[dict] = []
            for sheet_name in combine_names:
                if sheet_name in wb.sheetnames:
                    combined.extend(_parse_excel_sheet(wb[sheet_name]))
            if combined:
                return combined

        # Try preferred single sheets for the given module
        preferred = _MODULE_PREFERRED_SHEETS.get(module, [])
        for sheet_name in preferred:
            if sheet_name in wb.sheetnames:
                rows = _parse_excel_sheet(wb[sheet_name])
                if rows:
                    return rows

        # Fall back to active sheet
        return _parse_excel_sheet(wb.active)
    except Exception as exc:
        raise ValueError(f"Error reading Excel sheet: {exc}") from exc


def _normalise(row: dict, *keys: str) -> str:
    """Return first matching value from row by trying multiple key variants."""
    lowered = {k.lower().strip(): v for k, v in row.items()}
    for key in keys:
        val = lowered.get(key.lower())
        if val is not None:
            return str(val).strip()
    return ""


# ── Step 3: Bulk upload sites ─────────────────────────────────────────────────

@router.post(
    "/step3/bulk",
    summary="Bulk upload sites from Excel/CSV",
    status_code=status.HTTP_202_ACCEPTED,
)
async def bulk_upload_sites(
    file: UploadFile = File(...),
    user: Annotated[CurrentUser, Depends(get_current_user)] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    content = await file.read()
    rows = _parse_file(content, file.filename or "upload.csv")
    sites = [
        {
            "name":                _normalise(r, "site name", "name", "site"),
            "type":                _normalise(r, "type", "site type") or "Site",
            "address":             _normalise(r, "address"),
            "postcode":            _normalise(r, "postcode", "post code", "zip"),
            "city":                _normalise(r, "city", "town"),
            "operationalStatus":   _normalise(r, "operational status", "operational_status", "status") or "Active",
            "workingStations":     _normalise(r, "number of working stations", "working stations", "workstations", "number_of_working_stations"),
            "capacity":            _normalise(r, "capacity"),
            "primaryProducts":     _normalise(r, "primary products", "primary_products", "products"),
            "hazardClassification": _normalise(r, "hazard classification", "hazard_classification", "hazard level", "hazard"),
        }
        for r in rows
        if _normalise(r, "site name", "name", "site")
    ]
    svc = OrgSetupService(db)
    result = svc.bulk_upload_sites(user, {"sites": sites})
    db.commit()
    return accepted(result, f"Uploaded {result['count']} sites")


# ── Step 4: Bulk upload users ─────────────────────────────────────────────────

@router.post(
    "/step4/bulk",
    summary="Bulk upload users from Excel/CSV",
    status_code=status.HTTP_202_ACCEPTED,
)
async def bulk_upload_users(
    file: UploadFile = File(...),
    user: Annotated[CurrentUser, Depends(get_current_user)] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    content = await file.read()
    rows = _parse_file(content, file.filename or "upload.csv")
    users = [
        {
            "name":       _normalise(r, "name", "full name", "employee name"),
            "email":      _normalise(r, "email", "email address"),
            "role":       _normalise(r, "role", "job title", "position"),
            "department": _normalise(r, "department", "dept", "team"),
        }
        for r in rows
        if _normalise(r, "email", "email address")
    ]
    svc = OrgSetupService(db)
    result = svc.bulk_upload_users(user, {"users": users})
    db.commit()
    return accepted(result, f"Uploaded {result['count']} users")


# ── Step 4: HRMS import ───────────────────────────────────────────────────────

class HrmsImportPayload(BaseModel):
    url: str
    token: str = ""


@router.post(
    "/step4/hrms-import",
    summary="Import employees from HRMS API",
    status_code=status.HTTP_202_ACCEPTED,
)
async def hrms_import(
    payload: HrmsImportPayload = Body(...),
    user: Annotated[CurrentUser, Depends(get_current_user)] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    headers = {}
    if payload.token:
        headers["Authorization"] = f"Bearer {payload.token}"
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(payload.url, headers=headers)
            resp.raise_for_status()
            raw = resp.json()
    except httpx.HTTPError as exc:
        return accepted({"count": 0, "error": str(exc)}, "HRMS connection failed")

    # Try to extract a list from common HRMS response shapes
    employees: list = []
    if isinstance(raw, list):
        employees = raw
    elif isinstance(raw, dict):
        for key in ("data", "employees", "items", "results", "records"):
            if isinstance(raw.get(key), list):
                employees = raw[key]
                break

    users = [
        {
            "name":       str(emp.get("name") or emp.get("full_name") or emp.get("fullName") or ""),
            "email":      str(emp.get("email") or emp.get("work_email") or emp.get("workEmail") or ""),
            "role":       str(emp.get("role") or emp.get("job_title") or emp.get("jobTitle") or emp.get("designation") or ""),
            "department": str(emp.get("department") or emp.get("dept") or ""),
        }
        for emp in employees
        if emp.get("email") or emp.get("work_email") or emp.get("workEmail")
    ]

    svc = OrgSetupService(db)
    result = svc.bulk_upload_users(user, {"users": users})
    db.commit()
    return accepted(result, f"Imported {result['count']} employees from HRMS")


# ── Step 1: Excel parse ───────────────────────────────────────────────────────

_FIELD_MAP = {
    # Organisation ID
    "organisation id": "organisationId",
    "organization id": "organisationId",
    "org id": "organisationId",
    "org_id": "organisationId",
    # Organisation name
    "organisation name": "organisationName",
    "organization name": "organisationName",
    "org name": "organisationName",
    "company name": "organisationName",
    "company": "organisationName",
    "name": "organisationName",
    # Country
    "country": "country",
    # Industry sector
    "industry sector": "industrySector",
    "industry_sector": "industrySector",
    "industry type": "industrySector",
    "industry": "industrySector",
    "sector": "industrySector",
    # Number of employees
    "number of employees": "numberOfEmployees",
    "no. of employees": "numberOfEmployees",
    "no of employees": "numberOfEmployees",
    "employee count": "numberOfEmployees",
    "employees": "numberOfEmployees",
    "headcount": "numberOfEmployees",
    # Headquarters location
    "headquarters location": "headquartersLocation",
    "headquarters_location": "headquartersLocation",
    "headquarters": "headquartersLocation",
    "hq location": "headquartersLocation",
    "hq": "headquartersLocation",
    "headquarters address": "headquartersLocation",
    "hq address": "headquartersLocation",
    "head office": "headquartersLocation",
    # Parent company
    "parent company": "parentCompany",
    "parent_company": "parentCompany",
    "parent organisation": "parentCompany",
    "parent organization": "parentCompany",
    "parent": "parentCompany",
    # ISO 45001 status
    "iso 45001 status": "iso45001Status",
    "iso_45001_status": "iso45001Status",
    "iso45001 status": "iso45001Status",
    "iso 45001": "iso45001Status",
    "certification status": "iso45001Status",
    # Regulatory authority
    "regulatory authority": "regulatoryAuthority",
    "regulatory_authority": "regulatoryAuthority",
    "regulator": "regulatoryAuthority",
    "regulatory body": "regulatoryAuthority",
    # Establishment date
    "establishment date": "establishmentDate",
    "establishment_date": "establishmentDate",
    "founded": "establishmentDate",
    "date established": "establishmentDate",
    "incorporated": "establishmentDate",
}


@router.post("/step1/parse-excel", summary="Parse org details from Excel/CSV", status_code=200)
async def parse_org_excel(
    file: UploadFile = File(...),
    user: Annotated[CurrentUser, Depends(get_current_user)] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    content = await file.read()
    if not content:
        return {"success": False, "data": {"_error": "Uploaded file is empty"}}

    try:
        rows = _parse_file(content, file.filename or "upload.csv")
    except ValueError as exc:
        return {"success": False, "data": {"_error": str(exc)}}
    except Exception as exc:
        return {"success": False, "data": {"_error": f"Failed to parse file: {exc}"}}

    if not rows:
        return {"success": False, "data": {"_error": "No data rows found in file"}}

    result: dict = {}

    # Support two layouts:
    # 1. Key-value pairs: rows have "Field" and "Value" columns
    # 2. Header row with org field names as columns (single data row)
    if "field" in {k.lower().strip() for k in rows[0].keys()}:
        for row in rows:
            field = str(row.get("Field") or row.get("field") or "").lower().strip()
            value = str(row.get("Value") or row.get("value") or "").strip()
            mapped = _FIELD_MAP.get(field)
            if mapped and value:
                result[mapped] = value
    else:
        for row in rows[:1]:
            for col, val in row.items():
                mapped = _FIELD_MAP.get(col.lower().strip())
                if mapped and val:
                    result[mapped] = str(val).strip()

    if not result:
        sample_keys = list(rows[0].keys())[:10]
        return {
            "success": False,
            "data": {"_error": f"No recognised fields found. Column/field names in your file: {sample_keys}"},
        }

    return {"success": True, "data": result}


# ── Step 1: API connect ───────────────────────────────────────────────────────

class OrgApiConnectPayload(BaseModel):
    url: str
    api_key: str = ""
    token: str = ""


@router.post("/step1/api-connect", summary="Connect to external system and pull org data", status_code=200)
async def org_api_connect(
    payload: OrgApiConnectPayload = Body(...),
    user: Annotated[CurrentUser, Depends(get_current_user)] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    headers: dict[str, str] = {}
    if payload.token:
        headers["Authorization"] = f"Bearer {payload.token}"
    elif payload.api_key:
        headers["X-API-Key"] = payload.api_key

    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(payload.url, headers=headers)
            resp.raise_for_status()
            raw = resp.json()
    except httpx.HTTPError as exc:
        return {"success": False, "message": str(exc), "data": {}}

    # Normalise common API response shapes into org details
    org: dict = {}
    if isinstance(raw, dict):
        for key in ("org_id", "organisation_id", "organization_id", "orgId"):
            if raw.get(key):
                org["organisationId"] = str(raw[key])
                break
        for key in ("name", "organisation_name", "organization_name", "org_name", "company_name", "companyName"):
            if raw.get(key):
                org["organisationName"] = str(raw[key])
                break
        for key in ("country",):
            if raw.get(key):
                org["country"] = str(raw[key])
                break
        for key in ("industry_sector", "industrySector", "industry", "industry_type", "industryType", "sector"):
            if raw.get(key):
                org["industrySector"] = str(raw[key])
                break
        for key in ("number_of_employees", "numberOfEmployees", "employee_count", "employeeCount", "employees", "headcount"):
            if raw.get(key):
                org["numberOfEmployees"] = str(raw[key])
                break
        for key in ("headquarters_location", "headquartersLocation", "headquarters", "hq_location", "hq"):
            if raw.get(key):
                org["headquartersLocation"] = str(raw[key])
                break
        for key in ("parent_company", "parentCompany", "parent_organisation", "parent_organization"):
            if raw.get(key):
                org["parentCompany"] = str(raw[key])
                break
        for key in ("iso_45001_status", "iso45001Status", "iso45001", "certification_status"):
            if raw.get(key):
                org["iso45001Status"] = str(raw[key])
                break
        for key in ("regulatory_authority", "regulatoryAuthority", "regulator"):
            if raw.get(key):
                org["regulatoryAuthority"] = str(raw[key])
                break
        for key in ("establishment_date", "establishmentDate", "founded", "incorporated"):
            if raw.get(key):
                org["establishmentDate"] = str(raw[key])
                break

    return {"success": True, "data": org, "raw_preview": str(raw)[:500]}


# ── Step 1 template ───────────────────────────────────────────────────────────

@router.get("/step1/template", summary="Download org details CSV template")
def download_org_template():
    content = (
        "Field,Value\n"
        "Organisation ID,\n"
        "Organisation Name,\n"
        "Country,\n"
        "Industry Sector,\n"
        "Number of Employees,\n"
        "Headquarters Location,\n"
        "Parent Company,\n"
        "ISO 45001 Status,\n"
        "Regulatory Authority,\n"
        "Establishment Date,\n"
    )
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=org_details_template.csv"},
    )


# ── Template downloads ────────────────────────────────────────────────────────

@router.get("/step3/template", summary="Download sites CSV template")
def download_sites_template():
    content = (
        "Site Name,Address,Postcode,City,Type,Operational Status,Number of Working Stations,Capacity,Primary Products,Hazard Classification\n"
        "Bridgend Manufacturing Complex,\"Industrial Estate, Bridgend\",CF31 3TR,South Wales,Manufacturing & Assembly,Active,32,150,Wind Turbine Nacelles,High Risk\n"
        "Head Office,123 Corporate Blvd,SW1A 1AA,London,Site,Active,10,50,Administration,Low Risk\n"
        "Plant A,456 Industrial Ave,M1 2AB,Manchester,Plant,Active,20,100,Steel Components,Medium Risk\n"
    )
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sites_template.csv"},
    )


@router.get("/step4/template", summary="Download users CSV template")
def download_users_template():
    content = (
        "Name,Email,Role,Department\n"
        "John Smith,john.smith@company.com,HSE Manager,Safety\n"
        "Jane Doe,jane.doe@company.com,Supervisor,Operations\n"
        "Bob Wilson,bob.wilson@company.com,Site HSE Manager,Environment\n"
    )
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=users_template.csv"},
    )


# ── Generic module templates ──────────────────────────────────────────────────

_MODULE_TEMPLATES: dict[str, tuple[str, str]] = {
    "organisation": (
        "Organisation_ID,Organisation_Name,Country,Industry_Sector,Number_of_Employees,Headquarters_Location,Parent_Company,ISO_45001_Status,Regulatory_Authority,Establishment_Date\n"
        "ORG001,WindTech Nacelle Manufacturing Ltd,United Kingdom,Renewable Energy - Wind,150,Bridgend Wales,WindTech Group Plc,Certified,Health and Safety Executive (HSE),2015-03-15\n"
        "ORG002,Example Manufacturing Ltd,Germany,Automotive,320,Berlin,Example Group AG,In Progress,Bundesanstalt für Arbeitsschutz,2010-06-01\n",
        "organisation_template.csv",
    ),
    "sites": (
        "Site_ID,Site Name,Address,Postcode,City,Type,Operational_Status,Number_of_Working_Stations,Capacity,Primary_Products,Hazard_Classification\n"
        "SITE001,Bridgend Manufacturing Complex,\"Industrial Estate, Bridgend\",CF31 3TR,South Wales,Manufacturing & Assembly,Active,32,150,Wind Turbine Nacelles,High Risk\n"
        "SITE002,Head Office,123 Corporate Blvd,SW1A 1AA,London,Office,Active,10,50,Administration,Low Risk\n",
        "sites_template.csv",
    ),
    "departments": (
        "Department_ID,Site_ID,Department_Name,Manager_ID,Number_of_Teams\n"
        "1,SITE001,Heavy Assembly,EMP001,3\n"
        "2,SITE001,Maintenance,EMP002,2\n",
        "departments_template.csv",
    ),
    "working_stations": (
        "Station_ID,Station_Name,Site_ID,Department,Zone_Classification,Primary_Hazards,Staffing_Requirement,Equipment_List,Permit_Types_Required,Access_Restrictions\n"
        "STN001,Heavy Assembly Station 1,SITE001,Heavy Assembly,Heavy Assembly,HAZ001,3,Equipment Set 1,Hot Work,Authorized Personnel Only\n"
        "STN002,Welding Bay,SITE001,Heavy Assembly,Hot Work Zone,HAZ002,2,Equipment Set 2,Hot Work,PPE Required\n",
        "working_stations_template.csv",
    ),
    "roles": (
        "Role_ID,Role_Name,Job_Category,Authority_Level,Permit_Authority,Safety_Signatory\n"
        "ROLE001,Plant Manager,Senior Management,5,Yes,Yes\n"
        "ROLE002,HSE Manager,Safety,4,Yes,Yes\n",
        "roles_template.csv",
    ),
    "employees": (
        "Employee ID,Full Name,Date_of_Birth,Gender,Employment_Type,Employment_Start_Date,Job Title,Department,Shift_Pattern,Manager_ID,Induction_Date,Active_Status\n"
        "EMP001,Jessica Hernandez,1965-06-06,F,Permanent,2020-11-09,ROLE001,DEPT001,Rotating,,2020-11-21,Active\n"
        "EMP002,David Chen,1980-03-14,M,Permanent,2019-05-01,ROLE002,DEPT002,Days,EMP001,2019-05-15,Active\n",
        "employees_template.csv",
    ),
    "policies": (
        "Policy_ID,Policy_Name,Category,Issue_Date,Owner,Status\n"
        "POL001,Hot Work Safety,Hot Work,2022-01-15,Safety Manager,Current\n"
        "POL002,Working at Height,Height Safety,2022-03-01,HSE Manager,Current\n",
        "policies_template.csv",
    ),
    "permit_types": (
        "Permit_Type_ID,Permit_Type_Name,Risk_Level,Validity_Period_Hours,Concurrent_Limit\n"
        "PT001,Hot Work Permit,Critical,8,5\n"
        "PT002,Work at Height Permit,High,12,3\n",
        "permit_types_template.csv",
    ),
    "hazard_categories": (
        "Hazard_Category_ID,Category_Name,Description\n"
        "HC001,Mechanical,Moving machinery rotating equipment\n"
        "HC002,Chemical,Hazardous substances and chemical exposure\n",
        "hazard_categories_template.csv",
    ),
    "hazards": (
        "Hazard_ID,Category_ID,Hazard_Name,Severity,Probability\n"
        "HAZ001,HC001,Moving Machinery,Serious,Possible\n"
        "HAZ002,HC002,Chemical Exposure,Moderate,Unlikely\n",
        "hazards_template.csv",
    ),
    "training_programs": (
        "Training_ID,Training_Name,Duration_Hours,Frequency,Certification,Expiry_Months\n"
        "TRN001,Permit-to-Work System,4,Annual,Yes,12\n"
        "TRN002,Manual Handling,2,Biennial,No,24\n",
        "training_programs_template.csv",
    ),
    "permits_to_work": (
        "Permit_ID,Permit_Type_ID,Date_Issued,Time_Issued,Location_Station_ID,Work_Description,Duration_Requested_Hours,Issued_By,Approved_By,Validity_Start,Validity_End,Work_Start_Actual,Work_End_Actual,Number_of_Workers,Status,Deviation_Reported,Incident_Occurred\n"
        "PTW000001,PT001,2024-01-01,12:09,STN022,Welding,8,EMP036,EMP149,2024-01-01 16:09,2024-01-02 00:09,2024-01-01 16:09,,4,Active,Yes,No\n"
        "PTW000002,PT002,2024-01-02,08:00,STN005,Roof inspection,6,EMP010,EMP020,2024-01-02 09:00,2024-01-02 15:00,2024-01-02 09:00,2024-01-02 14:30,2,Closed,No,No\n",
        "permits_to_work_template.csv",
    ),
    "incidents": (
        "Incident_ID,Report_Date,Incident_DateTime,Location_Station,Incident_Type,Severity,Number_Persons_Involved,Description,Immediate_Cause,Root_Cause,Hazard_Involved,Permit_Active,Control_Failure,Reported_By,Investigation_Status,CAPA_Generated,Days_Away,Root_Cause_Category\n"
        "INC00001,2024-04-28,2024-04-28 15:37,STN031,Damage,Minor,3,Incident description,Human Error,Training Deficiency,HAZ001,,Yes,EMP020,In Progress,Yes,0,Training\n"
        "INC00002,2024-05-10,2024-05-10 09:15,STN012,Injury,Serious,1,Worker struck by falling object,Inadequate guarding,Engineering Deficiency,HAZ003,PTW000001,Yes,EMP045,Closed,Yes,3,Engineering\n",
        "incidents_template.csv",
    ),
    "near_misses": (
        "Near_Miss_ID,Report_Date,Event_DateTime,Location_Station,Description,Potential_Consequence,Hazard_Involved,Underlying_Cause,Control_Failure,Reported_By,CAPA_Escalation\n"
        "NM00001,2024-03-09,2024-03-09 09:39,STN019,Near-miss description,Environmental Release,HAZ010,Procedure Gap,No,EMP057,Yes\n"
        "NM00002,2024-04-01,2024-04-01 14:22,STN003,Tool dropped from height,Injury,HAZ001,Inadequate training,Yes,EMP023,No\n",
        "near_misses_template.csv",
    ),
    "safety_walks": (
        "Inspection_ID,Inspection_DateTime,Location_Station,Inspector,Inspection_Type,Issues_Found,Critical_Issues,Housekeeping_Rating,Compliance_Rating,Follow_Up_Required\n"
        "SW00001,2025-09-22 09:50,STN020,EMP053,Routine,2,1,2,2,No\n"
        "SW00002,2025-10-05 11:00,STN015,EMP031,Planned,0,0,4,5,No\n",
        "safety_walks_template.csv",
    ),
    "capa_actions": (
        "Action_ID,Incident_ID,Action_Type,Description,Root_Cause_Addressed,Responsible_Person,Due_Date,Status,Effectiveness_Rating\n"
        "CAPA00001,INC00001,Corrective,CAPA action for INC00001,Training,EMP037,2024-05-18,Completed,4\n"
        "CAPA00002,INC00002,Preventive,Install additional machine guarding,Engineering Deficiency,EMP012,2024-06-01,Open,\n",
        "capa_actions_template.csv",
    ),
    "shift_schedule": (
        "Schedule_ID,Employee_ID,Shift_Date,Shift_Type,Shift_Start,Shift_End,Actual_Hours_Worked,Station_Assigned,Supervisor\n"
        "SH00000001,EMP001,2024-01-01,Days,06:00,14:30,8.5,STN015,EMP118\n"
        "SH00000002,EMP002,2024-01-01,Nights,22:00,06:30,8.5,STN003,EMP118\n",
        "shift_schedule_template.csv",
    ),
    # ── legacy key aliases (same content, old key names used by UI) ──
    "incident_records": (
        "Incident_ID,Report_Date,Incident_DateTime,Location_Station,Incident_Type,Severity,Number_Persons_Involved,Description,Immediate_Cause,Root_Cause,Hazard_Involved,Permit_Active,Control_Failure,Reported_By,Investigation_Status,CAPA_Generated,Days_Away,Root_Cause_Category\n"
        "INC00001,2024-04-28,2024-04-28 15:37,STN031,Damage,Minor,3,Incident description,Human Error,Training Deficiency,HAZ001,,Yes,EMP020,In Progress,Yes,0,Training\n"
        "INC00002,2024-05-10,2024-05-10 09:15,STN012,Injury,Serious,1,Worker struck by falling object,Inadequate guarding,Engineering Deficiency,HAZ003,PTW000001,Yes,EMP045,Closed,Yes,3,Engineering\n",
        "incident_records_template.csv",
    ),
    "near_miss": (
        "Near_Miss_ID,Report_Date,Event_DateTime,Location_Station,Description,Potential_Consequence,Hazard_Involved,Underlying_Cause,Control_Failure,Reported_By,CAPA_Escalation\n"
        "NM00001,2024-03-09,2024-03-09 09:39,STN019,Near-miss description,Environmental Release,HAZ010,Procedure Gap,No,EMP057,Yes\n"
        "NM00002,2024-04-01,2024-04-01 14:22,STN003,Tool dropped from height,Injury,HAZ001,Inadequate training,Yes,EMP023,No\n",
        "near_miss_template.csv",
    ),
    "permit_records": (
        "Permit_ID,Permit_Type_ID,Date_Issued,Time_Issued,Location_Station_ID,Work_Description,Duration_Requested_Hours,Issued_By,Approved_By,Validity_Start,Validity_End,Work_Start_Actual,Work_End_Actual,Number_of_Workers,Status,Deviation_Reported,Incident_Occurred\n"
        "PTW000001,PT001,2024-01-01,12:09,STN022,Welding,8,EMP036,EMP149,2024-01-01 16:09,2024-01-02 00:09,2024-01-01 16:09,,4,Active,Yes,No\n"
        "PTW000002,PT002,2024-01-02,08:00,STN005,Roof inspection,6,EMP010,EMP020,2024-01-02 09:00,2024-01-02 15:00,2024-01-02 09:00,2024-01-02 14:30,2,Closed,No,No\n",
        "permit_records_template.csv",
    ),
    "sops_policies": (
        "Policy_ID,Policy_Name,Category,Issue_Date,Owner,Status\n"
        "POL001,Hot Work Safety,Hot Work,2022-01-15,Safety Manager,Current\n"
        "POL002,Working at Height,Height Safety,2022-03-01,HSE Manager,Current\n",
        "sops_policies_template.csv",
    ),
    "capa_data": (
        "Action_ID,Incident_ID,Action_Type,Description,Root_Cause_Addressed,Responsible_Person,Due_Date,Status,Effectiveness_Rating\n"
        "CAPA00001,INC00001,Corrective,CAPA action for INC00001,Training,EMP037,2024-05-18,Completed,4\n"
        "CAPA00002,INC00002,Preventive,Install additional machine guarding,Engineering Deficiency,EMP012,2024-06-01,Open,\n",
        "capa_data_template.csv",
    ),
    "hr_shift_data": (
        "Schedule_ID,Employee_ID,Shift_Date,Shift_Type,Shift_Start,Shift_End,Actual_Hours_Worked,Station_Assigned,Supervisor\n"
        "SH00000001,EMP001,2024-01-01,Days,06:00,14:30,8.5,STN015,EMP118\n"
        "SH00000002,EMP002,2024-01-01,Nights,22:00,06:30,8.5,STN003,EMP118\n",
        "hr_shift_data_template.csv",
    ),
    "training_records": (
        "Training_ID,Training_Name,Duration_Hours,Frequency,Certification,Expiry_Months\n"
        "TRN001,Permit-to-Work System,4,Annual,Yes,12\n"
        "TRN002,Manual Handling,2,Biennial,No,24\n",
        "training_records_template.csv",
    ),
    # ── kept for backward-compat but not aliased ──
    "audits": (
        "Audit Title,Audit Type,Standard,Scheduled Date,Lead Auditor,Status,Site\n"
        "Q1 Fire Safety Audit,Internal,ISO 45001,2024-03-15,Jane Doe,Scheduled,Bridgend Complex\n"
        "Annual External Audit,External,OSHA,2024-06-01,External Auditor,Scheduled,All Sites\n",
        "audits_template.csv",
    ),
    "audit_reports": (
        "Audit Title,Audit Type,Scheduled Date,Lead Auditor,Status\n"
        "Q1 Fire Safety Audit,Internal,2024-03-15,Jane Doe,Completed\n",
        "audit_reports_template.csv",
    ),
    "risk": (
        "Hazard Description,Location,Risk Level,Likelihood (1-5),Consequence (1-5),Controls,Responsible Person\n"
        "Machinery Contact/Crushing,STN001 Heavy Assembly,High,4,4,Machine guards fitted safety signs posted,Safety Manager\n"
        "Chemical Spill,Zone 4 Chemical Store,Critical,3,5,COSHH training spill kits secondary containment,EHS Lead\n",
        "risk_template.csv",
    ),
    "risk_assessments": (
        "Hazard Description,Location,Risk Level,Likelihood (1-5),Consequence (1-5),Controls,Responsible Person\n"
        "Machinery Contact,STN001,High,4,4,Machine guards fitted,Safety Manager\n",
        "risk_assessments_template.csv",
    ),
    "kpis": (
        "KPI Name,Period (Month),Value,Target,Unit,Site\n"
        "Incident Rate (TRIR),2024-01,2.4,3.0,per 100k hours,All Sites\n"
        "Near Miss Reports,2024-01,12,10,count,Bridgend Complex\n",
        "kpis_template.csv",
    ),
    "vendors": (
        "Company Name,Contact Email,Contact Phone,Service Type,HSE Rating,Contract Start,Contract End\n"
        "SafeWork Contractors Ltd,contact@safework.com,+44 1234 567890,Construction,Approved,2024-01-01,2024-12-31\n"
        "CleanTech Services,info@cleantech.com,+44 9876 543210,Cleaning,Conditional,2024-02-01,2024-07-31\n",
        "vendors_template.csv",
    ),
    "contractor_records": (
        "Company Name,Contact Email,Contact Phone,Service Type,HSE Rating,Contract Start,Contract End\n"
        "SafeWork Contractors Ltd,contact@safework.com,+44 1234 567890,Construction,Approved,2024-01-01,2024-12-31\n",
        "contractor_records_template.csv",
    ),
    "capa": (
        "Action_ID,Incident_ID,Action_Type,Description,Root_Cause_Addressed,Responsible_Person,Due_Date,Status,Effectiveness_Rating\n"
        "CAPA00001,INC00001,Corrective,CAPA action for INC00001,Training,EMP037,2024-05-18,Completed,4\n"
        "CAPA00002,INC00002,Preventive,Install additional machine guarding,Engineering Deficiency,EMP012,2024-06-01,Open,\n",
        "capa_template.csv",
    ),
    "permits": (
        "Permit_ID,Permit_Type_ID,Date_Issued,Time_Issued,Location_Station_ID,Work_Description,Duration_Requested_Hours,Issued_By,Approved_By,Validity_Start,Validity_End,Work_Start_Actual,Work_End_Actual,Number_of_Workers,Status,Deviation_Reported,Incident_Occurred\n"
        "PTW000001,PT001,2024-01-01,12:09,STN022,Welding,8,EMP036,EMP149,2024-01-01 16:09,2024-01-02 00:09,2024-01-01 16:09,,4,Active,Yes,No\n"
        "PTW000002,PT002,2024-01-02,08:00,STN005,Roof inspection,6,EMP010,EMP020,2024-01-02 09:00,2024-01-02 15:00,2024-01-02 09:00,2024-01-02 14:30,2,Closed,No,No\n",
        "permits_template.csv",
    ),
    "training": (
        "Training_ID,Training_Name,Duration_Hours,Frequency,Certification,Expiry_Months\n"
        "TRN001,Permit-to-Work System,4,Annual,Yes,12\n"
        "TRN002,Manual Handling,2,Biennial,No,24\n",
        "training_template.csv",
    ),
    "compliance": (
        "Policy_ID,Policy_Name,Category,Issue_Date,Owner,Status\n"
        "POL001,Hot Work Safety,Hot Work,2022-01-15,Safety Manager,Current\n"
        "POL002,Working at Height,Height Safety,2022-03-01,HSE Manager,Current\n",
        "compliance_template.csv",
    ),
    "users": (
        "Employee ID,Full Name,Date_of_Birth,Gender,Employment_Type,Employment_Start_Date,Job Title,Department,Shift_Pattern,Manager_ID,Induction_Date,Active_Status\n"
        "EMP001,Jessica Hernandez,1965-06-06,F,Permanent,2020-11-09,ROLE001,DEPT001,Rotating,,2020-11-21,Active\n"
        "EMP002,David Chen,1980-03-14,M,Permanent,2019-05-01,ROLE002,DEPT002,Days,EMP001,2019-05-15,Active\n",
        "users_template.csv",
    ),
}


@router.get("/template/{module_key}", summary="Download CSV template for any module")
def download_module_template(module_key: str):
    entry = _MODULE_TEMPLATES.get(module_key)
    if not entry:
        content, filename = "Field,Value\n", f"{module_key}_template.csv"
    else:
        content, filename = entry

    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Generic onboarding bulk import ────────────────────────────────────────────

@router.post(
    "/onboarding-bulk",
    summary="Bulk import any module during onboarding",
    status_code=status.HTTP_202_ACCEPTED,
)
async def onboarding_bulk_import(
    file: UploadFile = File(...),
    module: str = "incidents",
    user: Annotated[CurrentUser, Depends(get_current_user)] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    content = await file.read()
    try:
        rows = _parse_file(content, file.filename or "upload.csv", module=module)
    except Exception as exc:
        return accepted({"count": 0, "error": str(exc)}, "Parse failed")

    svc = OrgSetupService(db)
    result = svc.bulk_import_module(user, module, rows)
    db.commit()
    return accepted(result, f"Imported {result.get('count', 0)} records for {module}")
